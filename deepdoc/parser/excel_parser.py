#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#

import logging
import re
import sys
from io import BytesIO
from timeit import default_timer as timer

import pandas as pd
from openpyxl import Workbook, load_workbook
from PIL import Image

from rag.nlp import find_codec

# copied from `/openpyxl/cell/cell.py`
ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")


class RAGFlowExcelParser:
    @staticmethod
    def _load_excel_to_workbook(file_like_object):
        if isinstance(file_like_object, bytes):
            file_like_object = BytesIO(file_like_object)

        # Read first 4 bytes to determine file type
        file_like_object.seek(0)
        file_head = file_like_object.read(4)
        file_like_object.seek(0)

        if not (file_head.startswith(b"PK\x03\x04") or file_head.startswith(b"\xd0\xcf\x11\xe0")):
            logging.info("Not an Excel file, converting CSV to Excel Workbook")

            try:
                file_like_object.seek(0)
                df = pd.read_csv(file_like_object, on_bad_lines='skip')
                return RAGFlowExcelParser._dataframe_to_workbook(df)

            except Exception as e_csv:
                raise Exception(f"Failed to parse CSV and convert to Excel Workbook: {e_csv}")

        try:
            return load_workbook(file_like_object, data_only=True)
        except Exception as e:
            logging.info(f"openpyxl load error: {e}, try pandas instead")
            try:
                file_like_object.seek(0)
                try:
                    dfs = pd.read_excel(file_like_object, sheet_name=None)
                    return RAGFlowExcelParser._dataframe_to_workbook(dfs)
                except Exception as ex:
                    logging.info(f"pandas with default engine load error: {ex}, try calamine instead")
                    file_like_object.seek(0)
                    df = pd.read_excel(file_like_object, engine="calamine")
                    return RAGFlowExcelParser._dataframe_to_workbook(df)
            except Exception as e_pandas:
                raise Exception(f"pandas.read_excel error: {e_pandas}, original openpyxl error: {e}")

    @staticmethod
    def _clean_dataframe(df):
        def clean_string(s):
            if isinstance(s, str):
                return ILLEGAL_CHARACTERS_RE.sub(" ", s)
            return s

        if isinstance(df, dict):
            # Handle case where df is a dict of DataFrames (multiple sheets)
            for sheet_name, sheet_df in df.items():
                if isinstance(sheet_df, pd.DataFrame):
                    df[sheet_name] = sheet_df.apply(lambda col: col.map(clean_string))
            return df
        elif isinstance(df, pd.DataFrame):
            # Handle case where df is a single DataFrame
            return df.apply(lambda col: col.map(clean_string))
        else:
            # Return as-is if not a DataFrame or dict
            return df

    @staticmethod
    def _dataframe_to_workbook(df):
        # Clean the dataframe first
        cleaned_df = RAGFlowExcelParser._clean_dataframe(df)

        # Handle dict case (multiple sheets)
        if isinstance(cleaned_df, dict):
            if len(cleaned_df) > 1:
                return RAGFlowExcelParser._dataframes_to_workbook(cleaned_df)
            else:
                # Handle single sheet case
                sheet_name = next(iter(cleaned_df.keys()))
                cleaned_df = cleaned_df[sheet_name]

        # Ensure cleaned_df is a DataFrame before accessing columns
        if not isinstance(cleaned_df, pd.DataFrame):
            raise Exception(f"Expected DataFrame, got {type(cleaned_df)}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        for col_num, column_name in enumerate(cleaned_df.columns, 1):
            ws.cell(row=1, column=col_num, value=column_name)

        for row_num, row in enumerate(cleaned_df.values, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        return wb

    @staticmethod
    def _dataframes_to_workbook(dfs: dict):
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        for sheet_name, df in dfs.items():
            df = RAGFlowExcelParser._clean_dataframe(df)
            ws = wb.create_sheet(title=sheet_name)
            for col_num, column_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_num, value=column_name)
            for row_num, row in enumerate(df.values, 2):
                for col_num, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
        return wb

    @staticmethod
    def _extract_images_from_worksheet(ws, sheetname=None):
        """
        Extract images from a worksheet and enrich them with vision-based descriptions.

        Returns: List[dict]
        """
        images = getattr(ws, "_images", [])
        if not images:
            return []

        raw_items = []

        for img in images:
            try:
                img_bytes = img._data()
                # 限制图片大小，避免处理过大的图片
                if len(img_bytes) > 10 * 1024 * 1024:  # 10MB
                    logging.warning(f"Image too large ({len(img_bytes)/1024/1024:.2f}MB), skipping")
                    continue
                pil_img = Image.open(BytesIO(img_bytes)).convert("RGB")

                # 调整图片大小，减少内存使用
                max_size = 1536
                if pil_img.width > max_size or pil_img.height > max_size:
                    pil_img.thumbnail((max_size, max_size), Image.Resampling.BILINEAR)

                anchor = img.anchor
                if hasattr(anchor, "_from") and hasattr(anchor, "_to"):
                    r1, c1 = anchor._from.row + 1, anchor._from.col + 1
                    r2, c2 = anchor._to.row + 1, anchor._to.col + 1
                    if r1 == r2 and c1 == c2:
                        span = "single_cell"
                    else:
                        span = "multi_cell"
                else:
                    r1, c1 = anchor._from.row + 1, anchor._from.col + 1
                    r2, c2 = r1, c1
                    span = "single_cell"

                item = {
                    "sheet": sheetname or ws.title,
                    "image": pil_img,
                    "image_description": "",
                    "row_from": r1,
                    "col_from": c1,
                    "row_to": r2,
                    "col_to": c2,
                    "span_type": span,
                }
                raw_items.append(item)
            except Exception as e:
                logging.warning(f"Error extracting image: {e}")
                continue
        return raw_items

    def html(self, fnm, chunk_rows=256):
        from html import escape

        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        wb = RAGFlowExcelParser._load_excel_to_workbook(file_like_object)
        tb_chunks = []

        def _fmt(v):
            if v is None:
                return ""
            return str(v).strip()

        MAX_COLS = 100
        CONSECUTIVE_EMPTY = 100  # 连续空行阈值

        for sheetname in wb.sheetnames:
            ws = wb[sheetname]

            # 获取表头（第一行）
            header_row = None
            for row in ws.iter_rows(min_row=1, max_row=1, max_col=MAX_COLS, values_only=True):
                header_row = row
                break

            if not header_row or not any(header_row):
                continue

            # 生成表头 HTML
            tb_rows_0 = "<tr>" + "".join(f"<th>{escape(_fmt(v))}</th>" for v in header_row) + "</tr>"

            # 流式读取数据行
            current_chunk_rows = []
            consecutive_empty = 0
            total_rows = 0

            for row in ws.iter_rows(min_row=2, max_col=MAX_COLS, values_only=True):
                # 空行检查
                if not any(cell for cell in row if cell):
                    consecutive_empty += 1
                    if consecutive_empty >= CONSECUTIVE_EMPTY:
                        break  # 连续100行空，结束
                    continue
                else:
                    consecutive_empty = 0

                # 格式化为 HTML 行
                row_html = "<tr>" + "".join(
                    "<td></td>" if v is None else f"<td>{escape(_fmt(v))}</td>"
                    for v in row
                ) + "</tr>"
                current_chunk_rows.append(row_html)
                total_rows += 1

                # 达到 chunk 大小，生成一个 table
                if len(current_chunk_rows) >= chunk_rows:
                    tb = f'<table><caption>{escape(sheetname)}</caption>{tb_rows_0}{"".join(current_chunk_rows)}</table>\n'
                    tb_chunks.append(tb)
                    current_chunk_rows = []

            # 处理剩余的 rows
            if current_chunk_rows:
                tb = f'<table><caption>{escape(sheetname)}</caption>{tb_rows_0}{"".join(current_chunk_rows)}</table>\n'
                tb_chunks.append(tb)

        return tb_chunks


    def markdown(self, fnm):
        import pandas as pd

        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        try:
            file_like_object.seek(0)
            df = pd.read_excel(file_like_object)
        except Exception as e:
            logging.warning(f"Parse spreadsheet error: {e}, trying to interpret as CSV file")
            file_like_object.seek(0)
            df = pd.read_csv(file_like_object, on_bad_lines='skip')
        df = df.replace(r"^\s*$", "", regex=True)
        return df.to_markdown(index=False)

    def __call__(self, fnm):
        file_obj = BytesIO(fnm) if not isinstance(fnm, str) else fnm

        # 优先尝试普通模式（小文件更快）
        try:
            st = timer()
            wb = RAGFlowExcelParser._load_excel_to_workbook(file_obj)
            logging.info(f"Excel workbook loaded in {timer() - st:.2f}s")
            logging.info(f"Number of sheets: {len(wb.sheetnames)}")

            # 快速检测：估算总单元格数
            total_cells = 0
            CELL_THRESHOLD = 100000  # 5万单元格阈值

            for sheet in wb.worksheets:
                # 用 dimensions 快速估算，不实际遍历
                total_cells += sheet.max_row * sheet.max_column
                if total_cells > CELL_THRESHOLD:
                    break

            logging.info(f"Number of total_cells: {total_cells}")
            # 小文件：普通模式（快）
            if total_cells <= CELL_THRESHOLD:
                return self._parse_normal(wb)
            else:
                # 大文件或异常：流式模式
                return self._parse_streaming(file_obj)

        except Exception:
            pass


    def _parse_normal(self, wb):
        """小文件快速模式"""
        res = []
        MAX_COLS = 100
        CONSECUTIVE_EMPTY = 100  # 连续空行阈值
        st = timer()
        for sheetname in wb.sheetnames:
            logging.info(f"_parse_normal sheet: {sheetname}")
            ws = wb[sheetname]
            # 限制列数，但行数不限（最多到 ws.max_row）
            max_col = min(ws.max_column, MAX_COLS)

            rows = []
            consecutive_empty = 0

            # 遍历所有行，直到连续空行阈值或实际最大行
            for row in ws.iter_rows(min_row=1, max_col=max_col, values_only=True):
                if any(cell for cell in row if cell):
                    rows.append(row)
                    consecutive_empty = 0
                else:
                    consecutive_empty += 1
                    if consecutive_empty >= CONSECUTIVE_EMPTY:
                        break  # 连续100行空，结束当前sheet

            if len(rows) < 2:
                continue

            header, *data = rows
            for row in data:
                line = self._format_row(header, row, ws.title)
                if line:
                    res.append(line)
        logging.info(f"_parse_normal Parsed in {timer() - st:.2f}s")
        logging.info(f"_parse_normal Total rows processed: {len(res)}")
        return res


    def _parse_streaming(self, file_obj):
        """大文件流式模式（极低内存）"""
        file_obj.seek(0)
        wb = load_workbook(file_obj, read_only=True, data_only=True)

        res = []
        CONSECUTIVE_EMPTY = 100
        st = timer()
        for sheetname in wb.sheetnames:
            logging.info(f"_parse_streaming sheet: {sheetname}")
            ws = wb[sheetname]
            row_iter = ws.iter_rows(values_only=True)

            try:
                header = next(row_iter)
            except StopIteration:
                continue

            consecutive_empty = 0

            for row in row_iter:
                if any(cell for cell in row if cell):
                    line = self._format_row(header, row, ws.title)
                    if line:
                        res.append(line)
                    consecutive_empty = 0
                else:
                    consecutive_empty += 1
                    if consecutive_empty >= CONSECUTIVE_EMPTY:
                        break  # 连续100行空，结束当前sheet
        logging.info(f"_parse_streaming Parsed in {timer() - st:.2f}s")
        logging.info(f"_parse_streaming Total rows processed: {len(res)}")
        return res



    def _format_row(self, header, row, sheet_name):
        """统一格式化"""
        fields = []
        for i, cell in enumerate(row):
            if not cell:
                continue
            h = str(header[i]) if i < len(header) and header[i] else ""
            fields.append(f"{h}：" + str(cell) if h else str(cell))

        line = "; ".join(fields)
        if sheet_name.lower().find("sheet") < 0:
            line += " ——" + sheet_name
        return line


    @staticmethod
    def row_number(fnm, binary):
        if fnm.split(".")[-1].lower().find("xls") >= 0:
            wb = RAGFlowExcelParser._load_excel_to_workbook(BytesIO(binary))
            total = 0

            for sheetname in wb.sheetnames:
                try:
                    ws = wb[sheetname]
                    # 快速估算：优先用 max_row，但做简单校验
                    estimated = ws.max_row

                    # 如果估算值过大（>10000），抽样检查实际有效行
                    if estimated > 10000:
                        actual = 0
                        empty_streak = 0
                        for row in ws.iter_rows(values_only=True):
                            if any(cell for cell in row if cell):
                                actual += 1
                                empty_streak = 0
                            else:
                                empty_streak += 1
                                if empty_streak >= 100:
                                    break  # 连续100行空，停止
                        total += actual
                    else:
                        total += estimated

                except Exception as e:
                    logging.warning(f"Skip sheet '{sheetname}' due to rows access error: {e}")
                    continue
            return total

        if fnm.split(".")[-1].lower() in ["csv", "txt"]:
            encoding = find_codec(binary)
            txt = binary.decode(encoding, errors="ignore")
            return len(txt.split("\n"))


if __name__ == "__main__":
    psr = RAGFlowExcelParser()
    psr(sys.argv[1])
